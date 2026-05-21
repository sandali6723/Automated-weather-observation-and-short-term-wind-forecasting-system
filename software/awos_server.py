"""
AWOS Flask Backend Server with Advanced ML Forecasting
Random Forest (Wind Speed) + SVR (Wind Direction)
File: awos_server.py
"""

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import json
import os
from datetime import datetime, timedelta
import threading
import time
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.svm import SVR
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__, static_folder='public')
CORS(app)

# Global data storage
latest_weather_data = {
    'utcDate': 'N/A',
    'utcTime': 'N/A',
    'localTime': 'N/A',
    'temperature': 0,
    'humidity': 0,
    'pressure': 0,
    'windSpeed': 0,
    'windDirection': 0,
    'dewPoint': 0,
    'voltage': 0,
    'current': 0,
    'power': 0,
    'powerStatus': 'N/A',
    'commMode': 'N/A',
    'latitude': 'N/A',
    'longitude': 'N/A',
    'timestamp': datetime.now().isoformat()
}

weather_history = []
MAX_HISTORY = 8640  # 30 days at 5-minute intervals
MAX_AGE_DAYS = 30
LOG_INTERVAL = 300  # 5 minutes in seconds
last_logged_time = 0
wind_forecast_data = None
last_forecast_update = 0
FORECAST_UPDATE_INTERVAL = 300  # 5 minutes

# ==================== SENSOR OFFSETS ====================
# Stored and applied in base units: degC, %, hPa, degC (dewPoint), m/s, degrees
OFFSETS_FILE = 'sensor_offsets.json'

DEFAULT_OFFSETS = {
    'temperature':   0.0,
    'humidity':      0.0,
    'pressure':      0.0,
    'dewPoint':      0.0,
    'windSpeed':     0.0,
    'windDirection': 0.0
}

sensor_offsets = dict(DEFAULT_OFFSETS)

def load_offsets_from_disk():
    """Load sensor offsets from JSON file on disk."""
    global sensor_offsets
    if os.path.exists(OFFSETS_FILE):
        try:
            with open(OFFSETS_FILE, 'r') as f:
                saved = json.load(f)
            sensor_offsets = {**DEFAULT_OFFSETS, **saved}
            print(f'📐 Loaded sensor offsets: {sensor_offsets}')
        except Exception as e:
            print(f'⚠️  Could not load offsets file: {e}')

def save_offsets_to_disk():
    """Persist sensor offsets to JSON file."""
    try:
        with open(OFFSETS_FILE, 'w') as f:
            json.dump(sensor_offsets, f, indent=2)
    except Exception as e:
        print(f'❌ Could not save offsets: {e}')

def apply_offsets(record):
    """Return a copy of record with sensor offsets applied (base units)."""
    r = dict(record)
    r['temperature']   = round(float(r.get('temperature',   0)) + sensor_offsets['temperature'],   2)
    r['humidity']      = round(float(r.get('humidity',      0)) + sensor_offsets['humidity'],      2)
    r['pressure']      = round(float(r.get('pressure',      0)) + sensor_offsets['pressure'],      2)
    r['dewPoint']      = round(float(r.get('dewPoint',      0)) + sensor_offsets['dewPoint'],      2)
    r['windSpeed']     = round(max(0.0, float(r.get('windSpeed', 0)) + sensor_offsets['windSpeed']), 3)
    r['windDirection'] = round((float(r.get('windDirection', 0)) + sensor_offsets['windDirection']) % 360, 1)
    return r

# ==================== UNIT CONVERSION FUNCTIONS ====================

def celsius_to_fahrenheit(celsius):
    """Convert Celsius to Fahrenheit"""
    return round(float(celsius) * 9/5 + 32, 1)

def ms_to_knots(ms):
    """Convert m/s to knots"""
    return round(float(ms) * 1.94384, 1)

def ms_to_mph(ms):
    """Convert m/s to miles per hour"""
    return round(float(ms) * 2.23694, 1)

def ms_to_kmh(ms):
    """Convert m/s to km/h"""
    return round(float(ms) * 3.6, 1)

def hpa_to_inhg(hpa):
    """Convert hPa to inches of mercury"""
    return round(float(hpa) * 0.02953, 2)

def hpa_to_kpa(hpa):
    """Convert hPa to kPa"""
    return round(float(hpa) / 10, 1)

# ==================== HELPER FUNCTIONS ====================

def calculate_dew_point(temp, humidity):
    """Calculate dew point using Magnus formula"""
    a = 17.27
    b = 237.7
    alpha = ((a * temp) / (b + temp)) + np.log(humidity / 100.0)
    dew_point = (b * alpha) / (a - alpha)
    return round(dew_point, 2)

def clean_old_data():
    """Remove data older than MAX_AGE_DAYS"""
    global weather_history
    now = datetime.now()
    max_age = timedelta(days=MAX_AGE_DAYS)
    
    initial_count = len(weather_history)
    weather_history = [
        record for record in weather_history
        if (now - datetime.fromisoformat(record['timestamp'])) <= max_age
    ]
    
    removed = initial_count - len(weather_history)
    if removed > 0:
        print(f"🗑️ FIFO Cleanup: Removed {removed} records older than {MAX_AGE_DAYS} days")
        print(f"📊 Remaining records: {len(weather_history)}")
    
    return removed

def should_log_data():
    """Check if enough time has passed to log data"""
    global last_logged_time
    now = time.time()
    if now - last_logged_time >= LOG_INTERVAL:
        last_logged_time = now
        return True
    return False

# ==================== ML FORECASTING ====================

def forecast_wind_speed_rf(wind_speeds, n_forecast=24, n_lags=48):
    """
    Forecast wind speed using Random Forest
    Returns: list of forecasted wind speeds (m/s)
    """
    try:
        if len(wind_speeds) < n_lags + 10:
            return None
        
        # Prepare training data
        def create_features(data, n_lags):
            X, y = [], []
            for i in range(n_lags, len(data)):
                X.append(data[i-n_lags:i])
                y.append(data[i])
            return np.array(X), np.array(y)
        
        X, y = create_features(wind_speeds, n_lags)
        
        # Train Random Forest
        rf_model = RandomForestRegressor(
            n_estimators=150,
            max_depth=20,
            min_samples_split=5,
            min_samples_leaf=2,
            random_state=42,
            n_jobs=-1
        )
        
        # Use recent data for training
        train_size = min(len(X), 100)
        rf_model.fit(X[-train_size:], y[-train_size:])
        
        # Multi-step forecast
        history = list(wind_speeds[-n_lags:])
        forecast = []
        
        for step in range(n_forecast):
            input_features = np.array(history[-n_lags:]).reshape(1, -1)
            prediction = rf_model.predict(input_features)[0]
            
            # Ensure realistic values (wind speed >= 0)
            prediction = max(0, prediction)
            
            forecast.append(float(prediction))
            history.append(prediction)
        
        return forecast
        
    except Exception as e:
        print(f"❌ Wind speed forecast error: {e}")
        return None

def forecast_wind_direction_svr(weather_data, n_forecast=24, n_lags=24):
    """
    Forecast wind direction using SVR (Support Vector Regression)
    Uses sin/cos transformation for circular data
    Returns: list of forecasted wind directions (degrees)
    """
    try:
        if len(weather_data) < n_lags + 20:
            print(f"⚠️ Need at least {n_lags + 20} records for SVR. Current: {len(weather_data)}")
            return None
        
        print(f"   🧭 Training SVR with {len(weather_data)} weather records...")
        
        # Extract features
        features_list = ['temperature', 'humidity', 'pressure', 'dewPoint', 'windDirection', 'windSpeed']
        
        # Build DataFrame
        data_dict = {
            'temperature': [r['temperature'] for r in weather_data],
            'humidity': [r['humidity'] for r in weather_data],
            'pressure': [r['pressure'] for r in weather_data],
            'dewPoint': [r['dewPoint'] for r in weather_data],
            'windDirection': [r['windDirection'] for r in weather_data],
            'windSpeed': [r['windSpeed'] for r in weather_data]
        }
        df = pd.DataFrame(data_dict)
        
        # Compute u, v components and sin/cos for wind direction
        ws = df['windSpeed'].values
        wd = np.deg2rad(df['windDirection'].values)
        
        df['u'] = ws * np.cos(wd)
        df['v'] = ws * np.sin(wd)
        
        angles = np.deg2rad(df['windDirection'].values)
        y_sin = np.sin(angles)
        y_cos = np.cos(angles)
        
        # Create sequences
        X_seq, y_sin_seq, y_cos_seq = [], [], []
        
        for i in range(n_lags, len(df) - n_forecast + 1):
            X_seq.append(df[features_list].iloc[i-n_lags:i].values.flatten())
            y_sin_seq.append(y_sin[i:i+n_forecast])
            y_cos_seq.append(y_cos[i:i+n_forecast])
        
        if len(X_seq) < 10:
            print(f"⚠️ Not enough sequences. Need at least 10, got {len(X_seq)}")
            return None
        
        X_seq = np.array(X_seq)
        y_sin_seq = np.array(y_sin_seq)
        y_cos_seq = np.array(y_cos_seq)
        
        # Scale features
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X_seq)
        
        # Train SVR models for each forecast horizon
        print("   🔄 Training SVR models for each horizon...")
        svr_sin_models = []
        svr_cos_models = []
        
        for h in range(n_forecast):
            svr_sin = SVR(C=50, gamma='scale', epsilon=0.05)
            svr_cos = SVR(C=50, gamma='scale', epsilon=0.05)
            
            svr_sin.fit(X_scaled, y_sin_seq[:, h])
            svr_cos.fit(X_scaled, y_cos_seq[:, h])
            
            svr_sin_models.append(svr_sin)
            svr_cos_models.append(svr_cos)
            
            if (h + 1) % 6 == 0:
                print(f"      Trained {h+1}/{n_forecast} horizon models...")
        
        print(f"   ✅ SVR training complete for all {n_forecast} horizons")
        
        # Generate forecast for the most recent sequence
        latest_seq = df[features_list].iloc[-n_lags:].values.flatten().reshape(1, -1)
        latest_scaled = scaler.transform(latest_seq)
        
        forecast_sin = np.array([model.predict(latest_scaled)[0] for model in svr_sin_models])
        forecast_cos = np.array([model.predict(latest_scaled)[0] for model in svr_cos_models])
        
        # Convert sin/cos back to angles
        forecast_angles = np.rad2deg(np.arctan2(forecast_sin, forecast_cos)) % 360
        forecast_angles = [float(angle) for angle in forecast_angles]
        
        print(f"   ✅ Wind direction forecast complete: {len(forecast_angles)} horizons")
        return forecast_angles
        
    except Exception as e:
        print(f"❌ Wind direction forecast error: {e}")
        return None

def generate_wind_forecast():
    """Generate wind speed and direction forecast"""
    if len(weather_history) < 48:
        return {
            'success': False,
            'error': f'Need at least 48 data points for forecast. Current: {len(weather_history)}',
            'dataPoints': len(weather_history)
        }
    
    print('🔮 Generating wind forecast...')
    
    # Extract wind speeds (in m/s)
    wind_speeds = [r['windSpeed'] for r in weather_history]
    
    # Forecast wind speed
    speed_forecast = forecast_wind_speed_rf(wind_speeds)
    
    if not speed_forecast:
        return {
            'success': False,
            'error': 'Wind speed forecast failed',
            'dataPoints': len(weather_history)
        }
    
    # Forecast wind direction
    direction_forecast = forecast_wind_direction_svr(weather_history)
    
    if not direction_forecast:
        print('⚠️ Direction forecast failed, using simple persistence')
        latest_direction = weather_history[-1]['windDirection']
        direction_forecast = [latest_direction] * len(speed_forecast)
    
    # Create forecast timeline (next 2 hours, 5-minute intervals)
    last_time = datetime.fromisoformat(weather_history[-1]['timestamp'])
    forecast_times = [(last_time + timedelta(minutes=5*(i+1))).strftime('%H:%M') 
                      for i in range(len(speed_forecast))]
    
    return {
        'success': True,
        'forecast': {
            'times': forecast_times,
            'windSpeed': speed_forecast,  # m/s
            'windDirection': direction_forecast
        },
        'dataPoints': len(weather_history),
        'forecastHorizon': '2 hours (24 intervals)',
        'models': {
            'windSpeed': 'Random Forest Regressor',
            'windDirection': 'Support Vector Regression (SVR)'
        }
    }

# ==================== API ENDPOINTS ====================

@app.route('/')
def serve_index():
    """Serve the main HTML page"""
    return send_from_directory('public', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    """Serve static files"""
    return send_from_directory('public', path)

@app.route('/api/weather/latest', methods=['GET'])
def get_latest():
    """Get latest weather data"""
    global latest_weather_data
    
    # Calculate data age
    if latest_weather_data.get('timestamp'):
        last_update = datetime.fromisoformat(latest_weather_data['timestamp'])
        data_age = (datetime.now() - last_update).total_seconds()
    else:
        data_age = 9999  # Large number if no data
    
    # Determine connection status
    if data_age < 30:
        connection_status = "Connected - Live data"
    elif data_age < 300:  # 5 minutes
        connection_status = f"Stale data ({int(data_age)}s old)"
    else:
        connection_status = "No recent data"
    
    response_data = dict(latest_weather_data)
    response_data['dataAge'] = data_age
    response_data['connectionStatus'] = connection_status
    
    return jsonify(response_data)

@app.route('/api/weather/history', methods=['GET'])
def get_history():
    """Get weather history. Records logged after an offset change already have the offset baked in."""
    return jsonify(weather_history)


@app.route('/api/weather/download-excel', methods=['POST'])
def download_excel():
    """Generate Excel and return as base64 for desktop app compatibility"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        import base64
        
        data = request.json
        selected_date = data.get('date')
        
        if not selected_date:
            return jsonify({'success': False, 'error': 'No date provided'}), 400
        
        print(f'📥 Download request for: {selected_date}')
        
        # Parse selected date
        try:
            year, month, day = map(int, selected_date.split('-'))
        except ValueError as e:
            return jsonify({'success': False, 'error': 'Invalid date format'}), 400
        
        # Filter data
        filtered_data = []
        
        for record in weather_history:
            utc_date = record.get('utcDate', '')
            if not utc_date or utc_date == 'N/A':
                continue
            
            try:
                if '/' in utc_date:
                    parts = utc_date.split('/')
                    if len(parts) == 3:
                        d, m, y = map(int, parts)
                        if y < 100:
                            y += 2000
                        record_date = (y, m, d)
                elif '-' in utc_date:
                    parts = utc_date.split('-')
                    if len(parts) == 3:
                        y, m, d = map(int, parts)
                        record_date = (y, m, d)
                else:
                    continue
                
                if record_date == (year, month, day):
                    filtered_data.append(record)
                    
            except (ValueError, IndexError):
                continue
        
        print(f'✅ Found {len(filtered_data)} records')
        
        if len(filtered_data) == 0:
            return jsonify({
                'success': False, 
                'error': f'No data found for {selected_date}'
            }), 404
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Weather Data'
        
        headers = [
            'UTC Date', 'UTC Time', 'Local Time', 'Temperature (°C)', 'Humidity (%)',
            'Pressure (hPa)', 'Wind Speed (kt)', 'Wind Direction (°)', 'Dew Point (°C)',
            'Voltage (V)', 'Current (A)', 'Power (W)', 'Power Status',
            'Communication Mode', 'Latitude', 'Longitude'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, record in enumerate(filtered_data, 2):
            wind_speed_kt = round(float(record.get('windSpeed', 0)) * 1.94384, 1)
            
            row_data = [
                record.get('utcDate', 'N/A'),
                record.get('utcTime', 'N/A'),
                record.get('localTime', 'N/A'),
                round(float(record.get('temperature', 0)), 1),
                round(float(record.get('humidity', 0)), 1),
                round(float(record.get('pressure', 0)), 1),
                wind_speed_kt,
                round(float(record.get('windDirection', 0)), 0),
                round(float(record.get('dewPoint', 0)), 1),
                round(float(record.get('voltage', 0)), 2),
                round(float(record.get('current', 0)), 2),
                round(float(record.get('power', 0)), 2),
                record.get('powerStatus', 'N/A'),
                record.get('commMode', 'N/A'),
                record.get('latitude', 'N/A'),
                record.get('longitude', 'N/A')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Set column widths
        column_widths = [12, 10, 10, 15, 12, 14, 16, 18, 15, 12, 12, 12, 15, 20, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convert to base64
        excel_data = output.getvalue()
        base64_data = base64.b64encode(excel_data).decode('utf-8')
        
        filename = f'AWOS_Weather_Data_{selected_date}.xlsx'
        
        print(f'📥 Generated: {filename} ({len(excel_data)} bytes, {len(filtered_data)} records)')
        
        # Return base64 data instead of file
        return jsonify({
            'success': True,
            'filename': filename,
            'data': base64_data,
            'records': len(filtered_data),
            'size': len(excel_data)
        })
        
    except Exception as e:
        print(f'❌ Excel error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ALTERNATIVE: Add a direct file save endpoint for desktop
@app.route('/api/weather/save-excel-desktop', methods=['POST'])
def save_excel_desktop():
    """Save Excel directly to Downloads folder (desktop app only)"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from pathlib import Path
        
        data = request.json
        selected_date = data.get('date')
        
        if not selected_date:
            return jsonify({'success': False, 'error': 'No date provided'}), 400
        
        # Get Downloads folder
        downloads_folder = str(Path.home() / "Downloads")
        
        # Parse and filter (same as above)
        year, month, day = map(int, selected_date.split('-'))
        
        filtered_data = []
        for record in weather_history:
            utc_date = record.get('utcDate', '')
            if not utc_date or utc_date == 'N/A':
                continue
            
            try:
                if '/' in utc_date:
                    d, m, y = map(int, utc_date.split('/'))
                    if y < 100: y += 2000
                    record_date = (y, m, d)
                elif '-' in utc_date:
                    y, m, d = map(int, utc_date.split('-'))
                    record_date = (y, m, d)
                else:
                    continue
                
                if record_date == (year, month, day):
                    filtered_data.append(record)
            except:
                continue
        
        if len(filtered_data) == 0:
            return jsonify({'success': False, 'error': 'No data found'}), 404
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Weather Data'
        
        headers = [
            'UTC Date', 'UTC Time', 'Local Time', 'Temperature (°C)', 'Humidity (%)',
            'Pressure (hPa)', 'Wind Speed (kt)', 'Wind Direction (°)', 'Dew Point (°C)',
            'Voltage (V)', 'Current (A)', 'Power (W)', 'Power Status',
            'Communication Mode', 'Latitude', 'Longitude'
        ]
        
        # Style and write headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, record in enumerate(filtered_data, 2):
            row_data = [
                record.get('utcDate', 'N/A'),
                record.get('utcTime', 'N/A'),
                record.get('localTime', 'N/A'),
                round(float(record.get('temperature', 0)), 1),
                round(float(record.get('humidity', 0)), 1),
                round(float(record.get('pressure', 0)), 1),
                round(float(record.get('windSpeed', 0)) * 1.94384, 1),
                round(float(record.get('windDirection', 0)), 0),
                round(float(record.get('dewPoint', 0)), 1),
                round(float(record.get('voltage', 0)), 2),
                round(float(record.get('current', 0)), 2),
                round(float(record.get('power', 0)), 2),
                record.get('powerStatus', 'N/A'),
                record.get('commMode', 'N/A'),
                record.get('latitude', 'N/A'),
                record.get('longitude', 'N/A')
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Set column widths
        column_widths = [12, 10, 10, 15, 12, 14, 16, 18, 15, 12, 12, 12, 15, 20, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Save directly to Downloads folder
        filename = f'AWOS_Weather_Data_{selected_date}.xlsx'
        filepath = os.path.join(downloads_folder, filename)
        
        wb.save(filepath)
        
        print(f'💾 Saved to: {filepath} ({len(filtered_data)} records)')
        
        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'records': len(filtered_data)
        })
        
    except Exception as e:
        print(f'❌ Save error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/api/offsets', methods=['GET'])
def get_offsets():
    """Return current sensor offsets"""
    return jsonify(sensor_offsets)

@app.route('/api/offsets', methods=['POST'])
def set_offsets():
    """Update sensor offsets and persist to disk"""
    global sensor_offsets
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400

    allowed = set(DEFAULT_OFFSETS.keys())
    for key in allowed:
        if key in data:
            try:
                val = float(data[key])
                # Clamp wind direction offset to -359..359
                if key == 'windDirection':
                    val = max(-359.0, min(359.0, val))
                sensor_offsets[key] = round(val, 3)
            except (ValueError, TypeError):
                pass  # Skip invalid values

    save_offsets_to_disk()
    print(f'📐 Sensor offsets updated: {sensor_offsets}')
    return jsonify({'success': True, 'offsets': sensor_offsets})

@app.route('/api/weather/forecast', methods=['GET'])
def get_forecast():
    """Get wind forecast"""
    global wind_forecast_data, last_forecast_update
    
    print('🔮 Client requested wind forecast')
    
    # Check if we need to update forecast
    now = time.time()
    if not wind_forecast_data or (now - last_forecast_update) > FORECAST_UPDATE_INTERVAL:
        print('🔄 Generating new forecast...')
        wind_forecast_data = generate_wind_forecast()
        last_forecast_update = now
    
    return jsonify(wind_forecast_data)

@app.route('/api/weather/update', methods=['POST'])
def update_weather():
    """Update weather data (from ESP32) - stores data in original units"""
    global latest_weather_data
    
    data = request.json
    
    # Helper function to safely convert to float
    def safe_float(value, default=0):
        try:
            if value is None or value == '' or value == 'N/A':
                return default
            return float(value)
        except (ValueError, TypeError):
            return default
    
    # Safely convert all numeric values
    temperature = safe_float(data.get('temperature'))
    humidity = safe_float(data.get('humidity'))
    pressure = safe_float(data.get('pressure'))
    wind_speed = safe_float(data.get('windSpeed'))
    wind_direction = safe_float(data.get('windDirection'))
    voltage = safe_float(data.get('voltage'))
    current = safe_float(data.get('current'))
    power = safe_float(data.get('power'))
    
    # Calculate dew point only if temperature and humidity are valid
    if temperature != 0 and humidity != 0:
        try:
            dew_point = safe_float(data.get('dewPoint')) if data.get('dewPoint') else calculate_dew_point(temperature, humidity)
        except:
            dew_point = 0
    else:
        dew_point = 0
    
    timestamp = datetime.now().isoformat()
    
    # Build raw record first
    raw_record = {
        'utcDate': data.get('utcDate', datetime.utcnow().date().isoformat()),
        'utcTime': data.get('utcTime', datetime.utcnow().time().strftime('%H:%M:%S')),
        'localTime': data.get('localTime', datetime.now().time().strftime('%H:%M:%S')),
        'temperature': temperature,
        'humidity': humidity,
        'pressure': pressure,
        'windSpeed': wind_speed,
        'windDirection': wind_direction,
        'dewPoint': dew_point,
        'voltage': voltage,
        'current': current,
        'power': power,
        'powerStatus': data.get('powerStatus', 'N/A'),
        'commMode': data.get('commMode', 'N/A'),
        'latitude': data.get('latitude', 'N/A'),
        'longitude': data.get('longitude', 'N/A'),
        'timestamp': timestamp
    }

    # Apply offsets at write time — offsets baked into latest and new history records.
    # Old history records are NOT touched, so past data stays as originally logged.
    latest_weather_data = apply_offsets(raw_record)
    
    # Log sensor status
    sensors_working = []
    sensors_failed = []
    
    if temperature != 0 or humidity != 0:
        sensors_working.append('Weather')
    else:
        sensors_failed.append('Weather')
    
    if data.get('latitude') != 'N/A' and data.get('latitude') is not None:
        sensors_working.append('GPS')
    else:
        sensors_failed.append('GPS')
    
    status_msg = f"✅ Working: {', '.join(sensors_working) if sensors_working else 'None'}"
    if sensors_failed:
        status_msg += f" | ⚠️ Failed: {', '.join(sensors_failed)}"
    
    print(status_msg)
    
    # Log to history at 5-minute intervals
    if should_log_data():
        weather_history.append(dict(latest_weather_data))
        print(f"📝 Logged data point #{len(weather_history)} at {latest_weather_data['localTime']}")
        
        # FIFO cleanup
        if len(weather_history) > MAX_HISTORY:
            removed = weather_history.pop(0)
            print(f"🗑️ FIFO: Removed oldest record from {removed['utcDate']} {removed['utcTime']}")
    
    return jsonify({
        'success': True,
        'message': 'Data received',
        'dataPoints': len(weather_history),
        'logged': should_log_data(),
        'sensorsWorking': sensors_working,
        'sensorsFailed': sensors_failed
    })

@app.route('/api/weather/stats', methods=['GET'])
def get_stats():
    """Get weather statistics - returns data in original units"""
    if len(weather_history) == 0:
        return jsonify({'error': 'No data available'})
    
    temps = [r['temperature'] for r in weather_history]
    humidities = [r['humidity'] for r in weather_history]
    pressures = [r['pressure'] for r in weather_history]
    wind_speeds = [r['windSpeed'] for r in weather_history]
    
    stats = {
        'temperature': {
            'min': round(min(temps), 1),
            'max': round(max(temps), 1),
            'avg': round(sum(temps) / len(temps), 1)
        },
        'humidity': {
            'min': round(min(humidities), 1),
            'max': round(max(humidities), 1),
            'avg': round(sum(humidities) / len(humidities), 1)
        },
        'pressure': {
            'min': round(min(pressures), 1),
            'max': round(max(pressures), 1),
            'avg': round(sum(pressures) / len(pressures), 1)
        },
        'windSpeed': {
            'min': round(min(wind_speeds), 1),
            'max': round(max(wind_speeds), 1),
            'avg': round(sum(wind_speeds) / len(wind_speeds), 1)
        },
        'dataPoints': len(weather_history),
        'dataFrequency': '5 minutes',
        'coverageDays': round(len(weather_history) * 5 / (24 * 60), 1)
    }
    
    print('📈 Client requested statistics')
    return jsonify(stats)

@app.route('/api/weather/history', methods=['DELETE'])
def delete_history():
    """Clear weather history"""
    global weather_history, wind_forecast_data
    count = len(weather_history)
    weather_history = []
    wind_forecast_data = None
    print(f'🗑️ Deleted {count} history records')
    return jsonify({'success': True, 'message': 'History cleared', 'deletedCount': count})

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    uptime = time.time() - start_time
    return jsonify({
        'status': 'ok',
        'uptime': int(uptime),
        'dataPoints': len(weather_history),
        'lastUpdate': latest_weather_data['timestamp'],
        'maxHistory': MAX_HISTORY,
        'logInterval': '5 minutes',
        'forecastAvailable': len(weather_history) >= 48,
        'modelType': 'Random Forest + SVR',
        'requiresPyTorch': False
    })

# ==================== AUTO-SAVE & CLEANUP ====================

def auto_save_data():
    """Auto-save weather history every 10 minutes"""
    while True:
        time.sleep(600)  # 10 minutes
        if weather_history:
            try:
                with open('weather-backup.json', 'w') as f:
                    json.dump(weather_history, f, indent=2)
                print(f"💾 Auto-saved {len(weather_history)} records at {datetime.now().strftime('%H:%M:%S')}")
            except Exception as e:
                print(f"❌ Auto-save failed: {e}")

def auto_cleanup_data():
    """Auto-cleanup old data every hour"""
    while True:
        time.sleep(3600)  # 1 hour
        print('🧹 Running scheduled FIFO cleanup...')
        removed = clean_old_data()
        
        if removed > 0:
            try:
                with open('weather-backup.json', 'w') as f:
                    json.dump(weather_history, f, indent=2)
                print('💾 Saved after cleanup')
            except Exception as e:
                print(f"❌ Save after cleanup failed: {e}")

# ==================== STARTUP ====================

start_time = time.time()

def init_server():
    """Initialize server and load backup data"""
    global weather_history, latest_weather_data, last_logged_time
    
    # Load sensor offsets
    load_offsets_from_disk()
    
    # Load backup if exists
    if os.path.exists('weather-backup.json'):
        try:
            with open('weather-backup.json', 'r') as f:
                data = json.load(f)
                if isinstance(data, list) and len(data) > 0:
                    weather_history = data
                    print(f'📂 Loaded {len(weather_history)} previous records')
                    
                    if weather_history:
                        latest_weather_data = weather_history[-1]
                        last_time = datetime.fromisoformat(latest_weather_data['timestamp'])
                        last_logged_time = last_time.timestamp()
        except Exception as e:
            print(f'ℹ️ Could not load backup: {e}')
    
    # Start background threads
    threading.Thread(target=auto_save_data, daemon=True).start()
    threading.Thread(target=auto_cleanup_data, daemon=True).start()
    
    print('\n╔═══════════════════════════════════════════════════════════╗')
    print('║   🌤️  AWOS Flask Backend - ML Wind Forecasting            ║')
    print('╚═══════════════════════════════════════════════════════════╝')
    print('\n🔡 Server running on: http://0.0.0.0:3000')
    print('🌐 Access GUI at: http://localhost:3000')
    print('\n🤖 ML Models:')
    print('   ✅ Random Forest (Wind Speed)')
    print('   ✅ SVR - Support Vector Regression (Wind Direction)')
    print('\n📊 API Endpoints:')
    print('   - GET  /api/weather/latest')
    print('   - GET  /api/weather/history')
    print('   - GET  /api/weather/forecast')
    print('   - GET  /api/weather/stats')
    print('   - POST /api/weather/update')
    print('   - GET  /api/health')
    print('   - GET  /api/offsets       (read sensor offsets)')
    print('   - POST /api/offsets       (update sensor offsets)')
    print('\n⏱️  Data Logging:')
    print(f'   - Interval: Every 5 minutes')
    print(f'   - Max Records: {MAX_HISTORY} (30 days)')
    print(f'   - Current Records: {len(weather_history)}')
    print(f'   - Storage Units: °C, m/s, hPa (original units)')
    forecast_status = "Yes" if len(weather_history) >= 48 else f"No (need {48 - len(weather_history)} more records)"
    print(f'   - Forecast Ready: {forecast_status}')
    print('\n✅ Server ready! Waiting for data...\n')

def run_server():
    """Run Flask server"""
    init_server()
    app.run(host='0.0.0.0', port=3000, debug=False, threaded=True)

if __name__ == '__main__':
    run_server()
